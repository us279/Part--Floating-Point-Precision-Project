import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

# Parameters
Re = 100.0
Lx, Ly = 1.0, 1.0
nx, ny = 100, 100
dx, dy = Lx / (nx - 1), Ly / (ny - 1)
T_final = 0.5


CFL = 0.5
u,v =  exact_solution(X,Y,0)
speed = np.sqrt(u**2 + v**2)
max_speed = np.max(speed)
dt = CFL * min(dx, dy) / max_speed
# Grid
x = np.linspace(0, Lx, nx, dtype=np.float64)
y = np.linspace(0, Ly, ny, dtype=np.float64)
X, Y = np.meshgrid(x, y)

# Calculate the appropriate dt using the CFL condition

file_path = '/Users/uday03/Desktop/Part IIB Project/code/Part--Floating-Point-Precision-Project/Figs/DF_results_test.xlsx'
test_name = input('What are you testing?')

# DataFrame to store errors and time
error_df = pd.DataFrame(columns=['Precision', 'Max Error U', 'Mean Error U', 'RMSE U', 'Max Error V', 'Mean Error V', 'RMSE V', 'Time Taken'])
#error_df = pd.DataFrame(columns=['Precision','Mean Error U', 'Mean Error V', 'Time Taken'])

def testing_fp_precision(precision):
    fp_ = {np.float64: 'float64', np.float32: 'float32', np.float16: 'float16'}.get(precision)

    from time import time
    def exact_solution(xg, yg, t):
        u = precision(3/4 - 1./(4 * (1 + np.exp((-4 * xg + 4 * yg - t) * Re / 32))))
        v = precision(3/4 + 1./(4 * (1 + np.exp((-4 * xg + 4 * yg - t) * Re / 32))))
        return u, v

    # Initialize u and v arrays using the exact solution
    u_numerical, v_numerical = exact_solution(precision(X), precision(Y), 0)

    # Temporary arrays to hold new time step calculations
    u_next = np.zeros_like(u_numerical)
    v_next = np.zeros_like(v_numerical)

    # Arrays to store derivatives from the previous time step
    prev_dudt = np.zeros_like(u_numerical)
    prev_dvdt = np.zeros_like(v_numerical)

    # Time integration loop for numerical solution
    t0 = time()

    for t in np.arange(dt, T_final + dt, dt):
        # Compute gradients
        dudx, dudy = np.gradient(u_numerical, dx, dy)
        dvdx, dvdy = np.gradient(v_numerical, dx, dy)
        d2udx2, _ = np.gradient(dudx, dx)
        _, d2udy2 = np.gradient(dudy, dy)
        d2vdx2, _ = np.gradient(dvdx, dx)
        _, d2vdy2 = np.gradient(dvdy, dy)

        # Convert gradients to specified precision
        dudx, dudy = precision(dudx), precision(dudy)
        dvdx, dvdy = precision(dvdx), precision(dvdy)
        d2udx2, d2udy2 = precision(d2udx2), precision(d2udy2)
        d2vdx2, d2vdy2 = precision(d2vdx2), precision(d2vdy2)
  

        # Update derivatives
        dudt = precision(-(u_numerical * dudx + v_numerical * dudy) + (d2udx2 + d2udy2) / Re)
        dvdt = precision(-(u_numerical * dvdx + v_numerical * dvdy) + (d2vdx2 + d2vdy2) / Re)

        # Time stepping schemes
        if t > dt:
            u_next = precision(u_numerical + dt * (1.5 * dudt - 0.5 * prev_dudt))
            v_next = precision(v_numerical + dt * (1.5 * dvdt - 0.5 * prev_dvdt))
        else:  # Euler method for the first step
            u_next = precision(u_numerical + dt * dudt)
            v_next = precision(v_numerical + dt * dvdt)

        u_exact, v_exact = exact_solution(precision(X), precision(Y), t)
        u_next[0, :] = precision(u_exact[0, :])
        u_next[:, 0] = precision(u_exact[:, 0])
        v_next[0, :]= precision(v_exact[0, :])
        v_next[:, 0] = precision(v_exact[:, 0])

        # Update previous time step derivatives
        prev_dudt, prev_dvdt = precision(dudt), precision(dvdt)

        # Swap arrays for the next step
        u_numerical, u_next = u_next, u_numerical
        v_numerical, v_next = v_next, v_numerical

    # Time taken
    time_taken = time() - t0

    # Compute errors
    max_error_u = np.max(np.abs(u_exact - u_numerical))
    mean_error_u = np.mean(np.abs(u_exact - u_numerical))
    rmse_u = np.sqrt(np.mean(np.abs(u_exact - u_numerical)**2))
    max_error_v = np.max(np.abs(v_exact - v_numerical))
    mean_error_v = np.mean(np.abs(v_exact - v_numerical))
    rmse_v = np.sqrt(np.mean(np.abs(v_exact - v_numerical)**2))

    # Store errors and time in DataFrame
    errors = {
        'Test': test_name,
        'CFL': CFL,
        'Precision': fp_,
        'Max Error U': max_error_u,
        'Mean Error U': mean_error_u,
        'RMSE U': rmse_u,
        'Max Error V': max_error_v,
        'Mean Error V': mean_error_v,
        'RMSE V': rmse_v,
        'Time Taken': time_taken
    }

    # Plotting

    print(u_numerical[1,1])


    fig, axs = plt.subplots(2, 3, figsize=(18, 12))
    
    # fig.suptitle('Plots for a floating point precision of ' + str(fp_) , fontsize=16)

    # # Plot exact u
    # cf1 = axs[0, 0].contourf(X, Y, u_exact,levels= 100, cmap='viridis')
    # plt.colorbar(cf1, ax=axs[0, 0], orientation='vertical', label='u')
    # axs[0, 0].set_title('Exact Solution for u')
    # axs[0, 0].set_xlabel('x')
    # axs[0, 0].set_ylabel('y')

    # # Plot exact v
    # cf2 = axs[0, 1].contourf(X, Y, v_exact,levels= 100, cmap='viridis')
    # plt.colorbar(cf2, ax=axs[0, 1], orientation='vertical', label='v')
    # axs[0, 1].set_title('Exact Solution for v')
    # axs[0, 1].set_xlabel('x')
    # axs[0, 1].set_ylabel('y')

    # # Plot numerical u
    # cf3 = axs[1, 0].contourf(X, Y, u_numerical,levels= 100, cmap='viridis')
    # plt.colorbar(cf3, ax=axs[1, 0], orientation='vertical', label='u')
    # axs[1, 0].set_title('Numerical Solution for u')
    # axs[1, 0].set_xlabel('x')
    # axs[1, 0].set_ylabel('y')

    # # Plot numerical v
    # cf4 = axs[1, 1].contourf(X, Y, v_numerical,levels= 100, cmap='viridis')
    # plt.colorbar(cf4, ax=axs[1, 1], orientation='vertical', label='v')
    # axs[1, 1].set_title('Numerical Solution for v')
    # axs[1, 1].set_xlabel('x')
    # axs[1, 1].set_ylabel('y')

    # # Plot difference between exact and numerical u
    # diff_u = np.abs(u_exact - u_numerical)
    # cf5 = axs[1, 2].contourf(X, Y, diff_u,levels= 100, cmap='magma')
    # plt.colorbar(cf5, ax=axs[1, 2], orientation='vertical', label='|u_exact - u_numerical|')
    # axs[1, 2].set_title('Absolute Difference (u)')
    # axs[1, 2].set_xlabel('x')
    # axs[1, 2].set_ylabel('y')

    # # Plot difference between exact and numerical v
    # diff_v = np.abs(v_exact - v_numerical)
    # cf6 = axs[0, 2].contourf(X, Y, diff_v,levels= 100, cmap='magma')
    # plt.colorbar(cf6, ax=axs[0, 2], orientation='vertical', label='|v_exact - v_numerical|')
    # axs[0, 2].set_title('Absolute Difference (v)')
    # axs[0, 2].set_xlabel('x')
    # axs[0, 2].set_ylabel('y')

    # plt.tight_layout()
    plt.show()

    return errors

# Run tests and store results
for fp in [np.float64, np.float32, np.float16]:
    error_data = testing_fp_precision(fp)
    error_df = error_df.append(error_data, ignore_index=True)


print(error_df)

sheet_name = 'Sheet1'  # Specify the sheet name you want to update


book = load_workbook(file_path)
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    error_df.to_excel(writer, sheet_name=sheet_name, startrow=writer.sheets[sheet_name].max_row, index=False, header=False)

